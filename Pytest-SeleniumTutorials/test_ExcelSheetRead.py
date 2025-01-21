import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
def test_dataDriven():

        # below code is read the excel
        path = "C://Nishant_Study//nopcommerce//nopcommerceAdmin//TestData//TestData.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        row=sheet.max_row
        col=sheet.max_column

        for r in range(1,row+1):
                for c in range(1,col+1):
                        print(sheet.cell(row=r,column=c).value)

        # print("---------------------------------selenium code start--------------------")
        # driver = webdriver.Chrome()
        # driver.maximize_window()
        # driver.delete_all_cookies()
        # driver.get('https://admin:admin@the-internet.herokuapp.com/')
        # driver.find_element(By.XPATH, "//li//a[contains(text(),'Form Authentication')]").click()
        # userName = driver.find_element(By.ID,"username")
        # password = driver.find_element(By.ID, "password")
        # submit_btn=driver.find_element(By.XPATH,"//button[@type='submit']")
