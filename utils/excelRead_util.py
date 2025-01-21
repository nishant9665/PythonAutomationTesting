import openpyxl
import pandas as pd


def readExcelSheet(file_Name):
    path = "C://Nishant_Study//nopcommerce//nopcommerceAdmin//TestData//"+file_Name
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row = sheet.max_row
    col = sheet.max_column
    for r in range(1, row + 1):
        for c in range(1, col + 1):
            return sheet.cell(row=r, column=c).value

def read_excel_data(file_name):
    path = "C://Nishant_Study//nopcommerce//nopcommerceAdmin//TestData//" + file_name
    df = pd.read_excel(path)
    return df